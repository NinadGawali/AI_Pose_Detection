"""
Data Manager Module
Handles Excel operations for student practice records and statistics
"""

import pandas as pd
import os
import hashlib
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment


class DataManager:
    """Manages student practice data and statistics in Excel"""

    def __init__(self, excel_path="practice_data.xlsx"):
        self.excel_path = excel_path
        self.users_file = "users.xlsx"
        self._initialize_files()

    # ── Helpers ─────────────────────────────────────────────────────────────

    def _hash(self, password: str) -> str:
        """SHA-256 hash of a password."""
        return hashlib.sha256(password.encode()).hexdigest()

    def _initialize_files(self):
        """Create Excel files with headers if they don't exist."""
        if not os.path.exists(self.excel_path):
            df = pd.DataFrame(columns=[
                'Student_ID', 'Student_Name', 'Pose_Type',
                'Accuracy', 'Date', 'Time', 'Duration_Seconds',
                'Rules_Passed', 'Rules_Failed'
            ])
            df.to_excel(self.excel_path, index=False)
            self._format_excel(self.excel_path)

        if not os.path.exists(self.users_file):
            users_df = pd.DataFrame({
                'Username':   ['student1',              'student2',              'teacher1'],
                'Password':   [self._hash('pass123'),   self._hash('pass123'),   self._hash('admin123')],
                'Role':       ['student',               'student',               'teacher'],
                'Full_Name':  ['Demo Student 1',        'Demo Student 2',        'Demo Teacher'],
                'Created_At': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')] * 3,
            })
            users_df.to_excel(self.users_file, index=False)
            self._format_excel(self.users_file)

    def _format_excel(self, filepath):
        """Apply blue header formatting and auto-column widths."""
        try:
            wb = load_workbook(filepath)
            ws = wb.active
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column in ws.columns:
                max_length = 0
                col_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
            wb.save(filepath)
        except Exception as e:
            print(f"Formatting error: {e}")

    # ── Auth ─────────────────────────────────────────────────────────────────

    def authenticate_user(self, username, password, role):
        """
        Authenticate user. Returns (success: bool, full_name: str | None).
        Accepts both legacy plaintext passwords and new hashed passwords
        so existing demo accounts keep working after upgrade.
        """
        try:
            users_df = pd.read_excel(self.users_file, dtype=str).fillna("")
            uname = username.strip().lower()
            r     = role.strip().lower()

            # Try hashed match first
            hashed_match = users_df[
                (users_df['Username'].str.lower() == uname) &
                (users_df['Password']             == self._hash(password)) &
                (users_df['Role'].str.lower()     == r)
            ]
            if not hashed_match.empty:
                return True, hashed_match.iloc[0]['Full_Name']

            # Fallback: plaintext match (legacy rows) — auto-upgrades on success
            plain_match = users_df[
                (users_df['Username'].str.lower() == uname) &
                (users_df['Password']             == password) &
                (users_df['Role'].str.lower()     == r)
            ]
            if not plain_match.empty:
                # Upgrade plaintext → hashed in-place
                idx = plain_match.index[0]
                users_df.at[idx, 'Password'] = self._hash(password)
                users_df.to_excel(self.users_file, index=False)
                self._format_excel(self.users_file)
                return True, plain_match.iloc[0]['Full_Name']

            return False, None
        except Exception as e:
            print(f"Authentication error: {e}")
            return False, None

    def username_exists(self, username):
        """Check if a username already exists (case-insensitive)."""
        try:
            users_df = pd.read_excel(self.users_file, dtype=str).fillna("")
            return username.strip().lower() in users_df['Username'].str.lower().values
        except Exception as e:
            print(f"Error checking username: {e}")
            return False

    def register_user(self, username, password, full_name, role):
        """
        Register a new user.
        Returns (success: bool, message: str).
        """
        try:
            if not username or not password or not full_name:
                return False, "All fields are required."
            if len(username.strip()) < 3:
                return False, "Username must be at least 3 characters."
            if " " in username.strip():
                return False, "Username must not contain spaces."
            if len(password) < 6:
                return False, "Password must be at least 6 characters."
            if role not in ('student', 'teacher'):
                return False, "Invalid role."
            if self.username_exists(username):
                return False, "Username already exists. Please choose another."

            users_df = pd.read_excel(self.users_file, dtype=str).fillna("")

            new_row = pd.DataFrame([{
                'Username':   username.strip().lower(),
                'Password':   self._hash(password),
                'Role':       role,
                'Full_Name':  full_name.strip(),
                'Created_At': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            }])

            users_df = pd.concat([users_df, new_row], ignore_index=True)
            users_df.to_excel(self.users_file, index=False)
            self._format_excel(self.users_file)

            return True, "Account created successfully!"
        except Exception as e:
            print(f"Registration error: {e}")
            return False, f"Registration failed: {str(e)}"

    def get_all_users(self, role=None):
        """Get all users, optionally filtered by role."""
        try:
            users_df = pd.read_excel(self.users_file, dtype=str).fillna("")
            if role:
                users_df = users_df[users_df['Role'] == role]
            return users_df[['Username', 'Full_Name', 'Role']].to_dict('records')
        except Exception as e:
            print(f"Error getting users: {e}")
            return []

    # ── Practice records ──────────────────────────────────────────────────────

    def save_practice_record(self, student_id, student_name, pose_type,
                             accuracy, duration, rules_passed=None, rules_failed=None):
        """Save a student's practice session record."""
        try:
            now = datetime.now()
            new_record = {
                'Student_ID':       student_id,
                'Student_Name':     student_name,
                'Pose_Type':        pose_type,
                'Accuracy':         round(accuracy, 2),
                'Date':             now.strftime('%Y-%m-%d'),
                'Time':             now.strftime('%H:%M:%S'),
                'Duration_Seconds': duration,
                'Rules_Passed':     rules_passed or '',
                'Rules_Failed':     rules_failed or '',
            }
            df = pd.read_excel(self.excel_path)
            df = pd.concat([df, pd.DataFrame([new_record])], ignore_index=True)
            df.to_excel(self.excel_path, index=False)
            self._format_excel(self.excel_path)
            return True
        except Exception as e:
            print(f"Error saving record: {e}")
            return False

    # ── Analytics ─────────────────────────────────────────────────────────────

    def get_student_stats(self, student_id):
        """Get statistics for a specific student."""
        try:
            df = pd.read_excel(self.excel_path)
            student_data = df[df['Student_ID'] == student_id]
            if student_data.empty:
                return None
            return {
                'total_sessions':  len(student_data),
                'avg_accuracy':    round(student_data['Accuracy'].mean(), 2),
                'max_accuracy':    round(student_data['Accuracy'].max(), 2),
                'min_accuracy':    round(student_data['Accuracy'].min(), 2),
                'recent_sessions': student_data.tail(10).to_dict('records'),
                'pose_breakdown':  student_data.groupby('Pose_Type').agg(
                                       {'Accuracy': ['mean', 'count']}
                                   ).round(2).to_dict(),
            }
        except Exception as e:
            print(f"Error getting student stats: {e}")
            return None

    def get_all_students_summary(self):
        """Get summary statistics for all students."""
        try:
            df = pd.read_excel(self.excel_path)
            if df.empty:
                return []
            summary = df.groupby(['Student_ID', 'Student_Name']).agg(
                {'Accuracy': ['mean', 'count', 'max', 'min'], 'Date': 'max'}
            ).round(2)
            summary.columns = ['Avg_Accuracy', 'Total_Sessions',
                                'Max_Accuracy', 'Min_Accuracy', 'Last_Practice']
            return summary.reset_index().to_dict('records')
        except Exception as e:
            print(f"Error getting summary: {e}")
            return []

    def get_pose_statistics(self):
        """Get overall statistics by pose type."""
        try:
            df = pd.read_excel(self.excel_path)
            if df.empty:
                return {}
            pose_stats = df.groupby('Pose_Type').agg(
                {'Accuracy': ['mean', 'count', 'max', 'min']}
            ).round(2)
            pose_stats.columns = ['Average', 'Sessions', 'Best', 'Worst']
            return pose_stats.to_dict('index')
        except Exception as e:
            print(f"Error getting pose stats: {e}")
            return {}

    def get_recent_activities(self, limit=20):
        """Get recent practice activities across all students."""
        try:
            df = pd.read_excel(self.excel_path)
            if df.empty:
                return []
            return (df.sort_values(['Date', 'Time'], ascending=False)
                      .head(limit)
                      .to_dict('records'))
        except Exception as e:
            print(f"Error getting recent activities: {e}")
            return []